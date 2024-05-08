import pandas as pd
from openpyxl import load_workbook

# 엑셀 파일 로드
file_path = 'output_excel.xlsx'
new_file_path = '정리1.xlsx'
wb = load_workbook(file_path)
writer = pd.ExcelWriter(new_file_path, engine='openpyxl')

# 각 시트를 순회하면서 검사
i = 0
for sheet_name in wb.sheetnames:
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 데이터프레임 크기 조정
        while len(df.columns) < 3:
            df[f"추가 열 {len(df.columns) + 1}"] = ""  # 필요한 열 추가
        while len(df) < 3:
            df.loc[len(df)] = [""] * len(df.columns)  # 필요한 행 추가

        # 데이터 정제
        first_column = df.iloc[:, 0].astype(str).apply(lambda x: ''.join(x.split()))
        second_column = df.iloc[:, 1].astype(str).apply(lambda x: ''.join(x.split()))
        third_column = df.iloc[:, 2].astype(str).apply(lambda x: ''.join(x.split()))

        # 키워드 검사
        if (any('예산' in s and '백만원' in s for s in first_column)) or \
           (any('예산' in s and '백만원' in s for s in second_column)) or \
           (any('예산' in s and '백만원' in s for s in third_column)):
            i += 1
            if(i == 72) :  # 별첨 
                 i += 1
            sn = "과제" + str(i)
            print(i)
            df.to_excel(writer, sheet_name=sn)
    except Exception as e:
        print(f"시트 '{sheet_name}' 처리 중 오류 발생: {e}")

# 모든 조건을 만족하는 시트 저장
writer.close()
